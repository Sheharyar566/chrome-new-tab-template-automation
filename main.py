import os
import json
import shutil
import openpyxl
import pyperclip
from PIL import Image
from backup import backup
from git import Repo, remote
from packager import packager
from functions import writeLog
from functions import exceptionLogger

os.chdir('F://work')

work = './'
log_file = './log.txt'
ext_record = '../extensions.xlsx'
packed_dir = 'F:\\Private\\packed'
local_src = './chrome-new-tab-template'
remote_src = 'https://github.com/Sheharyar566/chrome-new-tab-template.git'

images_src = r'C:\Users\Sheha\Desktop\assets\images\\'
icons_src = r'C:\Users\Sheha\Desktop\assets\icons\original.jpg'
thumbs_src = r'C:\Users\Sheha\Desktop\assets\thumbs\\'

###################################################################
project = input('Project Name: ')
project = project.rstrip().capitalize().replace(' ', '-')
project_name = project + ' New Tab Wallpapers'

####################################################################
print('Creating the log file')

log = open(log_file, 'w+')
writeLog('Started a new project: ' + project)

####################################################################
writeLog('Checking if the project already exists')

try:
    record = openpyxl.load_workbook(filename = ext_record)
except Exception as e:
    exceptionLogger(e)

sheet = record[record.sheetnames[0]]
bottom_row = sheet.max_row + 1

for row in range(1, bottom_row):
    if(project.lower() == sheet.cell(row = row, column = 1).value):
        writeLog('Project already exists')
        print('Project already exists\nDeleting the log file')

        log.close()
        os.remove(log_file)

        exit()

writeLog('Perfect! The project is unique...ish')

####################################################################
repo = work + project_name
images_dest = repo + '/public/assets/images/'
thumbs_dest = repo + '/public/assets/thumbs/'
icons_dest = repo + '/public/assets/icons/'

####################################################################
class Progress(remote.RemoteProgress):
    def update(self, op_code, cur_count, max_count=None, message=''):
        writeLog(self._cur_line)

writeLog('Cloning the master repo from ')

if not os.path.isdir(local_src):
    try:
        localRepo = Repo.clone_from(remote_src, local_src, branch='master', progress=Progress())
    except Exception as e:
        exceptionLogger(e)

try:
    localRepo = Repo.clone_from(local_src, repo, branch='master', progress=Progress())
except Exception as e:
    exceptionLogger(e)

writeLog('Repo cloned successfully!')

####################################################################
writeLog('Copying images to public/assets/images folder')

try:
    src_files = os.listdir(images_src)
    i = 0
    for image in src_files:
        full_file_name = images_src + image
        if os.path.isfile(full_file_name):
            dest_file_name = images_dest + str(i) + '.jpg'
            shutil.copy(full_file_name, dest_file_name)
            writeLog('Copied file ' + str(i + 1) + '/15')
            i = i + 1
except Exception as e:
    exceptionLogger(e)

writeLog('Files copied successfully!')

####################################################################
writeLog('Creating thumbnails from the original images')

basewidth = 300
i = 0

try:
    for image in os.listdir(images_src):
        dest_thumb_name = thumbs_dest + str(i) + '.jpg'
        img = Image.open(images_src + image)
        
        wpercent = basewidth / float(img.size[0])
        hsize = int(float(img.size[1]) * float(wpercent))

        size = basewidth, hsize

        img.thumbnail(size, Image.ANTIALIAS)
        img.save(dest_thumb_name, optimize=True)

        i = i + 1
except Exception as e:
    exceptionLogger(e)

writeLog('Thumbnails created successfully!')

####################################################################
writeLog('Creating icon files')

try:
    for i in range(1, 20):
        multiple = i * 8

        if(multiple == 16 or multiple == 24 or multiple == 32 or multiple == 48 or multiple == 64 or multiple == 128):
            img = Image.open(icons_src)
            
            size = multiple, multiple
            img.thumbnail(size, Image.ANTIALIAS)
            img.save(icons_dest + 'icon-' + str(multiple) + '.png', optimize=True)

            img.close()
except Exception as e:
    exceptionLogger(e)

writeLog('Icons created successfully!')

####################################################################
writeLog('Editing the manifest.json')

try:
    manifestJSON = repo + '/public/manifest.json'
    with open(manifestJSON, 'r') as file:
        data = json.load(file)
        data['name'] = project_name.replace('-', ' ')
        data['description'] = project_name

    os.remove(manifestJSON)
    with open(manifestJSON, 'w') as file:
        json.dump(data, file, indent=4)
except Exception as e:
    exceptionLogger(e)

writeLog('Manifest.json edited sucessfully!')

####################################################################
writeLog('Creating the zip file for extension')

packed_extension = packed_dir + '/' + project.lower()
os.mkdir(packed_extension)

try:
    shutil.make_archive(packed_extension + '/' + project.lower(), 'zip', os.path.join(repo, 'public'))
except Exception as e:
    exceptionLogger(e)

writeLog('Zip created successfully!')

####################################################################
writeLog('Creating the package files')

pyperclip.copy('F:\\Work\\' + project_name + '\\public')

try:
    packager(project.lower())
except Exception as e:
    exceptionLogger(e)

writeLog('Package files created successfully!')

####################################################################
writeLog('Creating the backup for the project')

try:
    backup(project.lower(), icons_dest, images_dest)
except Exception as e:
    exceptionLogger(e)

writeLog('Backup created successfully!')

####################################################################
writeLog('Inserting project into the record')

try:
    sheet.append((project.lower(), '1.0.0', False, False))
    record.save(ext_record)
except Exception as e:
    exceptionLogger(e)

writeLog('Record inserted successfully!')

####################################################################
writeLog('Removing project directory!')

try:
    os.rmdir(repo)
except Exception as e:
    exceptionLogger(e)

writeLog('Project directory removed successfully!')

####################################################################
writeLog('Extension created successfully!')
print('Extension created successfully!')
print('Deleting the log file')

log.close()
os.remove(log_file)

print('Log file deleted successfully!')